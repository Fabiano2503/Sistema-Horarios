import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import Habilidad
from collections import OrderedDict

def export_niveles_habilidad_to_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Niveles de Habilidades"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Obtener todas las habilidades
    habilidades = Habilidad.objects.all().order_by('nombre')
    habilidad_nombres = [h.nombre for h in habilidades]

    # Encabezados
    headers = ['Practicante'] + habilidad_nombres
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    # Procesar datos
    practicantes_data = OrderedDict()
    for nivel in queryset:
        practicante_nombre = str(nivel.practicante)
        if practicante_nombre not in practicantes_data:
            practicantes_data[practicante_nombre] = {h_nombre: "" for h_nombre in habilidad_nombres}
        practicantes_data[practicante_nombre][str(nivel.habilidad)] = nivel.puntaje

    # Escribir datos
    row_idx = 2
    for practicante, puntajes in practicantes_data.items():
        row_data = [practicante] + [puntajes[h_nombre] for h_nombre in habilidad_nombres]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = alignment_left
            cell.border = thin_border
        row_idx += 1

    # Ajustar ancho de columnas
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        adjusted_width = max_length + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Congelar encabezado
    ws.freeze_panes = "A2"

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
